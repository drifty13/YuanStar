from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
import re

import cv2
import numpy as np
import pytest
from openpyxl import load_workbook

from yuanstar.catalog import load_catalog
from yuanstar.domain import (
    DetectedStarItem,
    GameVersion,
    ImportBatch,
    InventorySummaryRow,
    Quality,
    StarKind,
    normalize_instance_rows,
)
from yuanstar.export_excel import export_workbook
from yuanstar.session import SessionState
from yuanstar.ui_contract import inventory_display_rows
from yuanstar.vision.contracts import AnalysisResult, ImageInput
from yuanstar.vision.hierarchical_order import apply_hierarchical_order
from yuanstar.vision.experience_recognizer import recognize_experience_stones
from yuanstar.vision.models import CardCandidate, PageClassification, RecognizedStar
from yuanstar.vision.ocr_engine import OcrText


def _item(image_id: str, name: str, *, equipped_state: str = "unequipped") -> DetectedStarItem:
    return DetectedStarItem(
        card_id=f"{image_id}:card",
        source_image=image_id,
        source_position="r1c1",
        page_type="main",
        final_name=name,
        final_level=60,
        final_quality=Quality.ORANGE,
        equipped_state=equipped_state,
        confidence=.95,
        is_complete_card=True,
    )


def _batch(current: int | None, capacity: int | None) -> ImportBatch:
    return ImportBatch(
        imported_at=datetime.now(timezone.utc),
        image_count=1,
        game_version=GameVersion.RU_YUAN,
        bag_current_count=current,
        bag_capacity=capacity,
        ocr_executed=True,
    )


def test_clear_then_success_atomically_replaces_old_batch_and_unknowns_missing_metadata() -> None:
    state = SessionState(load_catalog())
    state.account_name = "共号"
    state.rows = [
        InventorySummaryRow(kind=StarKind.MAIN, name="天府", level=1, quantity=1)
    ]
    state.bag_current_count = 137
    state.bag_capacity = 245
    state.bag_manual_fields = {"bag_current_count", "bag_capacity"}
    state.experience_quantities = {"橙星曜": 9, "紫星曜": 8, "白星曜": 7}
    state.experience_manual_fields = {"橙星曜", "紫星曜", "白星曜"}

    state.clear_uploaded_images()
    assert state.pending_full_batch_replacement
    assert state.bag_current_count is None
    assert state.rows == []
    assert state.experience_quantities == {"橙星曜": None, "紫星曜": None, "白星曜": None}
    assert state.apply_local_analysis(AnalysisResult(False, "failed")) == 0
    assert state.bag_current_count is None

    state.uploaded_images = [ImageInput(id="new", filename="new.png", content=b"new")]
    state.apply_local_analysis(AnalysisResult(
        True,
        "ok",
        items=[_item("new", "武曲")],
        import_batch=_batch(None, None),
        image_pools={"new": "main"},
        experience_resolution={},
    ))

    assert state.account_name == "共号"
    assert state.bag_current_count is None and state.bag_capacity is None
    assert state.experience_quantities == {"橙星曜": None, "紫星曜": None, "白星曜": None}
    assert not state.bag_manual_fields and not state.experience_manual_fields
    assert [row.name for row in state.rows] == ["武曲"]
    assert not state.pending_full_batch_replacement


def test_new_batch_203_250_replaces_old_reviews_and_experience() -> None:
    state = SessionState(load_catalog())
    state.detected_items = [_item("old", "天府")]
    state.recalculate_postprocess()
    state.overlap_audit = [{"old": True}]
    state.experience_quantities = {"橙星曜": 0, "紫星曜": 2, "白星曜": 13}
    state.clear_uploaded_images()
    state.uploaded_images = [ImageInput(id="new", filename="new.png", content=b"new")]

    state.apply_local_analysis(AnalysisResult(
        True,
        "ok",
        items=[_item("new", "武曲", equipped_state="equipped")],
        import_batch=_batch(203, 250),
        image_pools={"new": "main"},
        overlap_audit=[],
        experience_resolution={
            "橙星曜": {"value": 1},
            "紫星曜": {"value": 77},
            "白星曜": {"value": 14},
        },
    ))

    assert (state.bag_current_count, state.bag_capacity) == (203, 250)
    assert state.experience_quantities == {"橙星曜": 1, "紫星曜": 77, "白星曜": 14}
    assert {item.source_image for item in state.detected_items} == {"new"}
    assert not state.overlap_audit
    assert state.rows[0].equipped_state == "equipped"


def _card(card_id: str, index: int) -> CardCandidate:
    return CardCandidate(card_id, 0, index, (index * 20, 0, 20, 20), (0, 0, 0, 0), True, .99)


def _star(
    card_id: str,
    level: int,
    *,
    quality: str = "橙",
    warning: str | None = None,
) -> RecognizedStar:
    return RecognizedStar(
        card_id,
        "main",
        "天府",
        "天府",
        .99,
        str(level),
        level,
        .99,
        .99,
        False,
        [warning] if warning else [],
        direct_level=level,
        quality=quality,
    )


def test_authoritative_segment_validator_only_marks_later_rising_item_and_clears_stale() -> None:
    cards = [_card("a", 0), _card("b", 1), _card("c", 2)]
    stars = [_star("a", 60, warning="level_order_conflict"), _star("b", 40), _star("c", 50)]
    result = apply_hierarchical_order(cards, stars, {
        card.card_id: ("equipped", .9, "test", [])
        for card in cards
    })
    by_id = {star.card_id: star for star in result}
    assert "level_order_conflict" not in by_id["a"].warnings
    assert by_id["a"].level == 60 and by_id["b"].level == 40
    assert by_id["c"].level is None
    assert by_id["c"].warnings == ["hierarchical_level_order_conflict"]


def test_unknown_and_changed_equipped_or_quality_start_new_level_segments() -> None:
    cards = [_card(str(index), index) for index in range(5)]
    stars = [
        _star("0", 1),
        _star("1", 60),
        _star("2", 1, quality="紫"),
        _star("3", 60, quality="白"),
        _star("4", 60, quality="白"),
    ]
    states = ["equipped", "unequipped", "unequipped", "unequipped", "unknown"]
    result = apply_hierarchical_order(cards, stars, {
        str(index): (state, .9, "test", [])
        for index, state in enumerate(states)
    })
    assert all("hierarchical_level_order_conflict" not in star.warnings for star in result)
    assert [star.level for star in result] == [1, 60, 1, 60, 60]


class _LocalMarkerEngine:
    def __init__(self, values: dict[int, str]) -> None:
        self.values = values

    def recognize(self, image, single_line=True):
        marker = int(image[-5, image.shape[1] // 2, 0]) if image.size else -1
        value = self.values.get(marker)
        return [OcrText(value, .96)] if value else []


def _experience_image(
    icons: list[tuple[int, tuple[int, int, int], int]],
) -> np.ndarray:
    image = np.zeros((700, 500, 3), dtype=np.uint8)
    image[20:50, 20:160] = 203
    image[650:680, 300:490] = 250
    for x, colour, marker in icons:
        cv2.circle(image, (x, 330), 48, colour, -1)
        image[365:382, max(0, x - 10):min(500, x + 55)] = marker
    return image


def test_experience_missing_or_single_kind_uses_local_icon_binding_only() -> None:
    page = PageClassification("experience", .9, ["selected_tab_visual:experience"])
    missing_orange = recognize_experience_stones(
        _experience_image([
            (180, (220, 80, 130), 77),
            (340, (255, 190, 60), 14),
        ]),
        (0, 0, 500, 700),
        _LocalMarkerEngine({77: "77", 14: "14"}),
        page=page,
    )
    assert missing_orange.complete
    assert (
        missing_orange.orange_count,
        missing_orange.purple_count,
        missing_orange.white_count,
    ) == (0, 77, 14)

    only_white = recognize_experience_stones(
        _experience_image([(250, (255, 190, 60), 14)]),
        (0, 0, 500, 700),
        _LocalMarkerEngine({14: "14"}),
        page=page,
    )
    assert only_white.complete
    assert (
        only_white.orange_count,
        only_white.purple_count,
        only_white.white_count,
    ) == (0, 0, 14)


def test_experience_count_roi_ignores_global_numbers_and_clips_safely() -> None:
    page = PageClassification("experience", .9, ["selected_tab_visual:experience"])
    image = _experience_image([(455, (255, 190, 60), 14)])
    result = recognize_experience_stones(
        image,
        (0, 0, 500, 700),
        _LocalMarkerEngine({14: "14", 203: "203", 250: "250"}),
        page=page,
    )
    assert result.orange_count in {0, None}
    assert result.purple_count in {0, None}
    assert result.white_count not in {203, 250}


def test_web_occurrence_sort_and_export_share_order_group_counts_and_account(tmp_path: Path) -> None:
    catalog = load_catalog()
    source = [
        InventorySummaryRow(
            id="unequipped",
            kind=StarKind.MAIN,
            name="天府",
            level=60,
            quality=Quality.ORANGE,
            quantity=1,
            equipped_state="unequipped",
            source_image_index=1,
        ),
        InventorySummaryRow(
            id="equipped",
            kind=StarKind.MAIN,
            name="武曲",
            level=1,
            quality=Quality.ORANGE,
            quantity=1,
            equipped_state="equipped",
            source_image_index=2,
        ),
        InventorySummaryRow(
            id="three",
            kind=StarKind.SUPPORT,
            name="解神",
            level=20,
            quality=Quality.PURPLE,
            quantity=3,
            equipped_state="unequipped",
        ),
    ]
    web_rows = normalize_instance_rows(source, catalog.order_index)
    web_display = inventory_display_rows(web_rows)
    destination = tmp_path / "current.xlsx"
    export_workbook(
        destination,
        GameVersion.RU_YUAN,
        "共号★",
        source,
        [],
        _batch(5, 250),
        catalog.order_index,
    )

    workbook = load_workbook(destination, data_only=True)
    values = list(workbook["背包汇总"].iter_rows(min_row=2, values_only=True))
    assert [row[3] for row in values] == [row.name for row in web_rows]
    assert [row[6] or "" for row in values] == [
        display["group_quantity"] for display in web_display
    ]
    assert len(values) == 5
    assert all(row[1] == "共号★" for row in values)
    assert [row[6] for row in values[-3:]] == ["本组共 3 颗", None, None]


@pytest.mark.anyio
@pytest.mark.nicegui_main_file("tests/batch_sort_experience_ui_app.py")
async def test_experience_original_preview_ui_smoke(user) -> None:
    user.javascript_rules[re.compile(".*scrollIntoView.*")] = lambda _: None
    await user.open("/")
    user.find("人工核对").click()
    await user.should_see(marker="experience-original-preview")
    user.find(marker="experience-original-preview").click()
    await user.should_see("experience-current.png")
    await user.should_see("100%")
    await user.should_see(marker="full-viewer-zoom-in")
    await user.should_see(marker="full-viewer-zoom-out")
    await user.should_see("第 1 / 1 张")
    await user.should_see(marker="full-viewer-close")
