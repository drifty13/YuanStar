from datetime import datetime, timezone

from openpyxl import load_workbook

from yuanstar.domain import ExperienceResource, GameVersion, ImportBatch, InventorySummaryRow, Quality, StarKind
from yuanstar.export_excel import export_workbook
from yuanstar.catalog import load_catalog


def test_excel_export_has_required_sheets_and_headers(tmp_path) -> None:
    destination = tmp_path / "inventory.xlsx"
    export_workbook(
        destination,
        GameVersion.RU_YUAN,
        "共号",
        [InventorySummaryRow(kind=StarKind.MAIN, name="武曲", level=10, quality=Quality.PURPLE, quantity=2)],
        [ExperienceResource(name="紫星曜", quantity=4, experience_per_item=500)],
        ImportBatch(imported_at=datetime.now(timezone.utc), image_count=2, game_version=GameVersion.RU_YUAN, bag_current_count=2, bag_capacity=250),
        load_catalog().order_index,
    )
    workbook = load_workbook(destination)
    assert workbook.sheetnames == ["背包汇总", "经验星石", "导入与校验"]
    assert [cell.value for cell in workbook["背包汇总"][1]] == ["游戏版本", "游戏账号名称", "大类", "星石名称", "品质", "当前等级", "计划等级", "数量"]
    assert [cell.value for cell in workbook["经验星石"][1]] == ["游戏版本", "游戏账号名称", "名称", "数量", "单颗经验"]
    assert [cell.value for cell in workbook["导入与校验"][1]][:7] == ["导入时间", "上传图片数量", "背包当前数量", "背包容量", "主星与辅星汇总数量", "差值", "完整性状态"]
    assert workbook["背包汇总"].column_dimensions["A"].width >= 11
    assert workbook["背包汇总"].column_dimensions["B"].width >= 15
    assert workbook["背包汇总"].column_dimensions["D"].width >= 11


def test_excel_uses_normalized_catalog_order(tmp_path) -> None:
    destination = tmp_path / "sorted.xlsx"
    export_workbook(
        destination,
        GameVersion.RU_YUAN,
        "",
        [
            InventorySummaryRow(kind=StarKind.SUPPORT, name="解神", level=1, quantity=1),
            InventorySummaryRow(kind=StarKind.MAIN, name="破军", level=10, quantity=1),
            InventorySummaryRow(kind=StarKind.MAIN, name="天府", level=5, quantity=1),
            InventorySummaryRow(kind=StarKind.MAIN, name="天府", level=5, quantity=2),
        ],
        [],
        None,
        load_catalog().order_index,
    )
    workbook = load_workbook(destination, data_only=True)
    rows = list(workbook["背包汇总"].iter_rows(min_row=2, values_only=True))
    assert [(row[2], row[3], row[5], row[6], row[7]) for row in rows] == [
        ("主星", "破军", 10, 10, "本组共 1 颗"),
        ("主星", "天府", 5, 5, "本组共 3 颗"),
        ("主星", "天府", 5, 5, None),
        ("主星", "天府", 5, 5, None),
        ("辅星", "解神", 1, 1, "本组共 1 颗"),
    ]
