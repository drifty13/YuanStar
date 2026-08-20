from pathlib import Path

from openpyxl import load_workbook
import pytest

from yuanstar.accounts import AccountWorkspaceManager
from yuanstar.app import (
    ImportTaskState,
    accumulated_engine_initializations,
    inventory_viewport_alignment_script,
    is_import_mutation_locked,
    reset_import_task_transients,
)
from yuanstar.catalog import load_catalog
from yuanstar.domain import DetectedStarItem, GameVersion, ImportBatch, InventorySummaryRow, Quality, StarKind
from yuanstar.export_excel import export_workbook
from yuanstar.persistence import WorkspaceStore
from yuanstar.session import SessionState
from yuanstar.ui_contract import plan_display_rows
from yuanstar.vision.contracts import AnalysisResult, ImageInput


def row(level: int = 20, *, name: str = "天府") -> InventorySummaryRow:
    return InventorySummaryRow(kind=StarKind.MAIN, name=name, quality=Quality.ORANGE, level=level, quantity=1)


def test_plan_targets_are_level_only_and_follow_current_instances() -> None:
    state = SessionState(load_catalog())
    instance_id = state.add_row(row(20))
    assert state.plan_targets == {}
    assert state.plan_level(instance_id) == 20

    state.set_plan_level(instance_id, 60)
    assert state.plan_targets == {instance_id: 60}
    displayed = plan_display_rows(state.filtered_rows(), state.plan_targets)
    assert displayed[0]["level"] == 60
    assert displayed[0]["current_level"] == 20

    state.update_row(instance_id, row(50))
    assert state.plan_level(instance_id) == 60
    state.update_row(instance_id, row(60))
    assert state.plan_level(instance_id) == 60
    assert state.plan_targets == {instance_id: 60}
    state.delete_row(instance_id)
    assert state.plan_targets == {}


def test_plan_validation_correction_and_undo_are_atomic() -> None:
    state = SessionState(load_catalog())
    instance_id = state.add_row(row(40))
    with pytest.raises(ValueError, match="不能低于当前等级 40"):
        state.set_plan_level(instance_id, 30)
    assert state.plan_level(instance_id) == 40

    state.correct_current_and_plan_level(instance_id, 30)
    assert state.selected_row().level == 30
    assert state.plan_level(instance_id) == 30
    assert state.undo()
    assert state.rows[0].level == 40
    assert state.plan_level(instance_id) == 40
    assert state.redo()
    assert state.rows[0].level == 30
    assert state.plan_level(instance_id) == 30


def test_ocr_row_lowering_updates_detected_value_and_survives_recalculation_history() -> None:
    state = SessionState(load_catalog())
    state.detected_items = [DetectedStarItem(
        card_id="ocr-card", source_image="ocr-image", source_position="r1c1",
        final_name="天府", final_level=40, final_quality=Quality.ORANGE,
        is_complete_card=True,
    )]
    state.recalculate_postprocess()
    instance_id = state.rows[0].star_instance_id

    state.correct_current_and_plan_level(instance_id, 30)
    assert state.rows[0].level == 30
    assert state.plan_level(instance_id) == 30
    assert state.detected_items[0].final_level == 30
    assert state.detected_items[0].manual_override is True
    state.recalculate_postprocess()
    assert state.rows[0].level == 30

    assert state.undo()
    assert state.detected_items[0].final_level == 40
    assert state.rows[0].level == 40
    assert state.redo()
    assert state.detected_items[0].final_level == 30
    state.recalculate_postprocess()
    assert state.rows[0].level == 30


def test_plan_input_switch_contract_keeps_effective_value_and_blocks_unresolved_low_confirmation() -> None:
    source = Path("src/yuanstar/app.py").read_text(encoding="utf-8")
    assert 'target_input.on("blur", lambda _: apply_plan_level())' in source
    assert 'target_input.value = str(state.plan_level(plan_editing.star_instance_id))' in source
    assert 'review_view_state["plan_input_blocked"] = True' in source
    assert source.count('if review_view_state.get("plan_input_blocked"):') == 4
    assert 'low_level_confirmation.on("hide", lambda _: restore_input_to_effective(close=False))' in source


def test_engine_initialization_count_never_regresses_on_sparse_events() -> None:
    assert accumulated_engine_initializations(0, 1) == 1
    assert accumulated_engine_initializations(1, 0) == 1
    assert accumulated_engine_initializations(1, None) == 1
    assert accumulated_engine_initializations(1, 2) == 2


def test_new_ocr_task_resets_transient_progress_without_changing_accumulation_rule() -> None:
    task = ImportTaskState(
        status="failed", total_images=4, completed_images=3, current_image_index=3,
        current_filename="old.png", error_count=2, error_summary="old", engine_initializations=1,
    )
    reset_import_task_transients(task, total_images=2)
    assert (task.total_images, task.completed_images, task.current_image_index, task.current_filename) == (2, 0, None, None)
    assert (task.error_count, task.error_summary, task.engine_initializations) == (0, None, 0)
    assert accumulated_engine_initializations(task.engine_initializations, 1) == 1


def test_import_mutation_lock_covers_restore_point_window_and_running_task() -> None:
    assert is_import_mutation_locked(ocr_busy=True, task_status="idle")
    assert is_import_mutation_locked(ocr_busy=False, task_status="running")
    assert not is_import_mutation_locked(ocr_busy=False, task_status="succeeded")


def test_restore_point_writes_a_prepared_fixed_snapshot_not_later_state(tmp_path: Path) -> None:
    state = SessionState(load_catalog())
    state.add_row(row(20))
    store = WorkspaceStore(tmp_path / "workspace")
    prepared = store.prepare(state)
    state.update_row(state.rows[0].star_instance_id, row(50))
    point = store.create_restore_point(prepared)
    restored = WorkspaceStore(point).load(load_catalog()).state
    assert restored is not None
    assert restored.rows[0].level == 20


def test_one_time_counterpart_alignment_remains_independent_of_in_place_reloads() -> None:
    script = inventory_viewport_alignment_script(
        ".current-inventory-table", ".planned-inventory-table",
    )
    assert "sourceMiddle.scrollTop" in script
    assert "targetMiddle.scrollTop" in script
    assert "__yuanstarInventoryAlignToken" in script


def test_plan_save_status_and_account_busy_contracts_are_local_and_monotonic() -> None:
    source = Path("src/yuanstar/app.py").read_text(encoding="utf-8")
    assert 'element.set_text(status)' in source
    assert 'element.classes(add="text-negative text-weight-bold", remove="text-grey")' in source
    assert 'editor_section.refresh()' not in source[source.index("def refresh_after_plan_change"):source.index("def apply_plan_level")]
    assert 'set_ocr_busy(True)\n                await start_import(confirmed=True)' in source
    assert 'restore_active_account_selector()' in source
    assert 'set_ocr_busy(False)' in source
    assert '"结果仍在当前内存，但自动保存失败。"' in source


def test_plan_buttons_and_workspace_restart_are_account_isolated(tmp_path: Path) -> None:
    catalog = load_catalog()
    manager = AccountWorkspaceManager(tmp_path / "YuanStar")
    first = manager.load_current(catalog).account
    second = manager.create_account("账号 B", GameVersion.RU_YUAN, catalog)
    state_a = SessionState(catalog)
    instance_a = state_a.add_row(row(10))
    assert state_a.plan_level_60(instance_a)
    manager.workspace_store(first.account_id).save(state_a)

    state_b = SessionState(catalog)
    instance_b = state_b.add_row(row(20, name="紫微"))
    state_b.set_plan_level(instance_b, 40)
    manager.workspace_store(second.account_id).save(state_b)

    loaded_a = manager.load_account(first.account_id, catalog).load_result.state
    loaded_b = manager.load_account(second.account_id, catalog).load_result.state
    assert loaded_a.plan_targets == {instance_a: 60}
    assert loaded_b.plan_targets == {instance_b: 40}
    assert loaded_b.restore_plan_to_current(instance_b)
    assert loaded_b.plan_targets == {}
    assert loaded_a.reset_all_plan_targets()
    assert loaded_a.plan_targets == {}
    assert loaded_a.undo() and loaded_a.plan_level(instance_a) == 60


def test_restore_points_keep_latest_three_complete_workspace_copies(tmp_path: Path) -> None:
    state = SessionState(load_catalog())
    state.add_row(row(20))
    state.uploaded_images = [ImageInput(filename="old.png", content=b"old")]
    store = WorkspaceStore(tmp_path / "workspace")
    for _ in range(4):
        store.create_restore_point(state)
    points = store.restore_point_paths()
    assert len(points) == 3
    restored = WorkspaceStore(points[-1]).load(load_catalog()).state
    assert restored is not None
    assert restored.rows[0].level == 20
    assert restored.uploaded_images[0].content == b"old"


def test_successful_ocr_rebuilds_plan_with_new_instance_ids_and_undo_redo() -> None:
    catalog = load_catalog()
    state = SessionState(catalog)
    old_id = state.add_row(row(20))
    state.set_plan_level(old_id, 60)
    image = ImageInput(filename="new.png", content=b"new", id="new-image")
    state.uploaded_images = [image]
    result = AnalysisResult(
        True,
        "ok",
        items=[DetectedStarItem(
            card_id="new-card", source_image=image.id, source_position="r1c1",
            final_name="天府", final_level=30, final_quality=Quality.ORANGE,
            is_complete_card=True,
        )],
        import_batch=ImportBatch(image_count=1, game_version=GameVersion.RU_YUAN),
        image_pools={image.id: "main"},
    )
    state.apply_local_analysis(result, rebuild_inventory=True)
    assert len(state.rows) == 1
    assert state.rows[0].star_instance_id != old_id
    assert state.plan_targets == {}
    assert state.undo()
    assert state.rows[0].star_instance_id == old_id
    assert state.plan_level(old_id) == 60
    assert state.redo()
    assert len(state.rows) == 1 and state.rows[0].star_instance_id != old_id and state.plan_targets == {}


def test_excel_exports_current_and_plan_levels_without_changing_experience(tmp_path: Path) -> None:
    state = SessionState(load_catalog())
    instance_id = state.add_row(row(20))
    state.set_plan_level(instance_id, 60)
    destination = export_workbook(
        tmp_path / "inventory.xlsx", state.game_version, "账号", state.rows,
        state.experience_resources(), None, state.catalog.order_index, state.plan_targets,
    )
    workbook = load_workbook(destination, data_only=True)
    assert [cell.value for cell in workbook["背包汇总"][1]] == [
        "游戏版本", "游戏账号名称", "大类", "星石名称", "品质", "当前等级", "计划等级", "数量",
    ]
    assert list(workbook["背包汇总"].iter_rows(min_row=2, values_only=True))[0][5:7] == (20, 60)
    assert [cell.value for cell in workbook["经验星石"][1]] == ["游戏版本", "游戏账号名称", "名称", "数量", "单颗经验"]
