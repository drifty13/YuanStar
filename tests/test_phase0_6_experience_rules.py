from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from yuanstar.experience_calculator import (
    ExperienceCalculationError,
    InstanceExperiencePlan,
    feedable_experience_for_instances,
    feedable_experience_required,
    owned_experience,
    raw_experience_required,
    remaining_gap,
    requirement_as_purple_white,
    stage_6_24_requirement,
    summarize_experience_plan,
)
from yuanstar.experience_rules import (
    DEFAULT_EXPERIENCE_RULES_PATH,
    ExperienceRuleLoadError,
    cached_experience_rules,
    load_experience_rules,
)


def _write_rules(path: Path, intervals: list[tuple[int, int, str, int, int]] | None = None) -> None:
    workbook = Workbook()
    config = workbook.active
    config.title = "Codex配置"
    config.append(["配置键", "值"])
    for key, value in {
        "max_level": 60,
        "white_exp": 100,
        "purple_exp": 500,
        "orange_exp": 1000,
        "include_orange_in_owned_exp": True,
        "assume_current_bar_progress": 0,
        "include_breakthrough_materials": False,
        "stage_6_24_purple_yield": 3,
        "stage_6_24_stamina_cost": 10,
        "round_exp_to": 100,
    }.items():
        config.append([key, value])
    sheet = workbook.create_sheet("升级经验区间")
    sheet.append(["当前等级起", "当前等级止", "经验条规则", "首级经验", "递增步长"])
    for row in intervals or [(1, 60, "固定值", 100, 0)]:
        sheet.append(row)
    workbook.save(path)


def test_project_relative_default_rules_load_and_are_immutable() -> None:
    rules = load_experience_rules()
    project_root = Path(__file__).resolve().parents[1]
    assert DEFAULT_EXPERIENCE_RULES_PATH == (
        project_root / "resources" / "reference" / "YuanStar_Phase0_6A_经验星曜规则与逐级数据.xlsx"
    )
    assert rules.max_level == 60
    assert len(rules.level_exp) == 60
    assert rules.level_exp[1] == 100 and rules.level_exp[60] == 15000
    with pytest.raises(TypeError):
        rules.level_exp[1] = 1  # type: ignore[index]


def test_loader_expands_arithmetic_and_fixed_intervals(tmp_path: Path) -> None:
    path = tmp_path / "rules.xlsx"
    _write_rules(path, [(1, 30, "等差递增", 100, 100), (31, 60, "固定值", 4000, 0)])
    rules = load_experience_rules(path)
    assert rules.level_exp[1] == 100
    assert rules.level_exp[30] == 3000
    assert rules.level_exp[31] == rules.level_exp[60] == 4000


@pytest.mark.parametrize(
    ("intervals", "message"),
    [
        ([(1, 59, "固定值", 100, 0)], "缺少60级"),
        ([(1, 40, "固定值", 100, 0), (40, 60, "固定值", 100, 0)], "区间重叠"),
    ],
)
def test_loader_rejects_missing_or_overlapping_levels(tmp_path: Path, intervals, message: str) -> None:
    path = tmp_path / "rules.xlsx"
    _write_rules(path, intervals)
    with pytest.raises(ExperienceRuleLoadError, match=message):
        load_experience_rules(path)


def test_loader_rejects_missing_sheet_and_invalid_value(tmp_path: Path) -> None:
    path = tmp_path / "rules.xlsx"
    _write_rules(path)
    workbook = Workbook()
    workbook.save(path)
    with pytest.raises(ExperienceRuleLoadError, match="缺少“Codex配置”工作表"):
        load_experience_rules(path)

    _write_rules(path)
    workbook = __import__("openpyxl").load_workbook(path)
    workbook["Codex配置"][2][1].value = "sixty"
    workbook.save(path)
    with pytest.raises(ExperienceRuleLoadError, match="最高等级不是有效正整数"):
        load_experience_rules(path)


def test_cached_loader_reads_once_per_service_lifecycle(monkeypatch) -> None:
    import yuanstar.experience_rules as module

    cached_experience_rules.cache_clear()
    calls = 0
    original = module.load_experience_rules

    def counted(path):
        nonlocal calls
        calls += 1
        return original(path)

    monkeypatch.setattr(module, "load_experience_rules", counted)
    assert cached_experience_rules() is cached_experience_rules()
    assert calls == 1
    cached_experience_rules.cache_clear()


def test_confirmed_core_calculations_and_per_instance_rounding() -> None:
    rules = load_experience_rules()
    assert raw_experience_required(1, 10, rules) == 2700
    assert feedable_experience_required(1, 10, rules) == 2700
    assert requirement_as_purple_white(2700, rules).purple == 5
    assert requirement_as_purple_white(2700, rules).white == 2
    assert raw_experience_required(10, 20, rules) == 9550
    assert feedable_experience_required(10, 20, rules) == 9600
    assert raw_experience_required(50, 60, rules) == 118000
    assert raw_experience_required(1, 60, rules) == 256750
    assert feedable_experience_required(1, 60, rules) == 256800
    assert feedable_experience_for_instances(
        [InstanceExperiencePlan("one", 10, 20), InstanceExperiencePlan("two", 10, 20)], rules
    ) == (2, 19200)


def test_zero_reverse_inventory_and_stage_cases() -> None:
    rules = load_experience_rules()
    assert raw_experience_required(60, 60, rules) == raw_experience_required(60, 59, rules) == 0
    assert owned_experience(1, 2, 3, rules) == 2300
    assert remaining_gap(1600, 2300) == 0
    assert stage_6_24_requirement(1600, rules).runs == 2
    assert stage_6_24_requirement(1600, rules).stamina == 20
    assert stage_6_24_requirement(0, rules).runs == 0
    with pytest.raises(ExperienceCalculationError):
        raw_experience_required(0, 60, rules)


def test_summary_uses_full_inventory_but_leaves_unknown_gap_unavailable() -> None:
    rules = load_experience_rules()
    plans = [
        InstanceExperiencePlan("first", 10, 20),
        InstanceExperiencePlan("second", 10, 20),
        InstanceExperiencePlan("already-done", 20, 20),
    ]
    summary = summarize_experience_plan(
        plans, rules, {"橙星曜": 1, "紫星曜": 1, "白星曜": 1}
    )
    assert summary.planned_instance_count == 2
    assert summary.required.experience == 19200
    assert summary.remaining is not None and summary.remaining.experience == 17600
    assert summary.stage_6_24 is not None and summary.stage_6_24.runs == 12

    unknown = summarize_experience_plan(plans, rules, {"橙星曜": None, "紫星曜": 1, "白星曜": 1})
    assert unknown.required.experience == 19200
    assert unknown.remaining is None and unknown.stage_6_24 is None


def test_summary_skips_only_invalid_instance_and_reports_its_id() -> None:
    rules = load_experience_rules()
    summary = summarize_experience_plan(
        [InstanceExperiencePlan("valid", 1, 10), InstanceExperiencePlan("bad-id", 61, 60)],
        rules,
        {"橙星曜": 0, "紫星曜": 0, "白星曜": 0},
    )
    assert summary.planned_instance_count == 1
    assert summary.required.experience == 2700
    assert summary.calculation_warnings == ("实例 bad-id 的等级超出经验规则范围",)
