from __future__ import annotations

from datetime import datetime, timezone

from openpyxl import load_workbook
import pytest

from yuanstar.catalog import load_catalog
from yuanstar.domain import (
    DetectedStarItem,
    GameVersion,
    ImportBatch,
    InventorySummaryRow,
    Quality,
    StarKind,
)
from yuanstar.export_excel import export_workbook
from yuanstar.session import SessionState
from yuanstar.ui_contract import inventory_display_rows, review_counts
from yuanstar.vision.models import CardCandidate
from yuanstar.vision.pipeline import LocalOfflineVisionPipeline


def _row(
    row_id: str,
    name: str,
    level: int,
    quality: Quality,
    equipped_state: str,
    *,
    source_index: int,
    kind: StarKind = StarKind.MAIN,
) -> InventorySummaryRow:
    return InventorySummaryRow(
        id=row_id,
        kind=kind,
        name=name,
        level=level,
        quality=quality,
        quantity=1,
        equipped_state=equipped_state,
        source_image_index=source_index,
    )


def _sorting_state() -> SessionState:
    state = SessionState(load_catalog())
    state.rows = [
        _row("tianfu-60-purple", "天府", 60, Quality.PURPLE, "unknown", source_index=4),
        _row("wuqu-60-orange", "武曲", 60, Quality.ORANGE, "equipped", source_index=5),
        _row("tianfu-40-orange", "天府", 40, Quality.ORANGE, "unequipped", source_index=6),
        _row("tianfu-60-orange-equipped", "天府", 60, Quality.ORANGE, "equipped", source_index=2),
        _row("tianfu-60-orange-unequipped", "天府", 60, Quality.ORANGE, "unequipped", source_index=1),
        _row(
            "support-tianfu-60-orange",
            "天府",
            60,
            Quality.ORANGE,
            "equipped",
            source_index=7,
            kind=StarKind.SUPPORT,
        ),
    ]
    return state


def test_web_default_sort_and_group_ignore_equipped() -> None:
    state = _sorting_state()

    rows = state.filtered_rows()
    assert [(row.name, row.level, row.quality) for row in rows] == [
        ("天府", 60, Quality.ORANGE),
        ("天府", 60, Quality.ORANGE),
        ("天府", 60, Quality.PURPLE),
        ("武曲", 60, Quality.ORANGE),
        ("天府", 40, Quality.ORANGE),
        ("天府", 60, Quality.ORANGE),
    ]
    assert [row.id for row in rows[:2]] == [
        "tianfu-60-orange-unequipped",
        "tianfu-60-orange-equipped",
    ]
    display = inventory_display_rows(rows)
    assert [row["group_quantity"] for row in display] == [
        "本组共 2 颗",
        "",
        "本组共 1 颗",
        "本组共 1 颗",
        "本组共 1 颗",
        "本组共 1 颗",
    ]


def test_web_active_filter_sorts_name_then_level_then_quality() -> None:
    state = _sorting_state()
    state.set_filters("全部", "全部", "天府 武曲")

    rows = state.filtered_rows()
    assert [(row.name, row.level, row.quality) for row in rows] == [
        ("天府", 60, Quality.ORANGE),
        ("天府", 60, Quality.ORANGE),
        ("天府", 60, Quality.PURPLE),
        ("天府", 40, Quality.ORANGE),
        ("武曲", 60, Quality.ORANGE),
        ("天府", 60, Quality.ORANGE),
    ]
    display = inventory_display_rows(rows, aggregate_by_name=True)
    assert [row["group_quantity"] for row in display] == [
        "本组共 4 颗",
        "",
        "",
        "",
        "本组共 1 颗",
        "本组共 1 颗",
    ]


def test_excel_uses_default_category_sort_and_group_without_structure_change(
    tmp_path,
) -> None:
    state = _sorting_state()
    destination = tmp_path / "inventory.xlsx"

    export_workbook(
        destination,
        GameVersion.RU_YUAN,
        "测试账号",
        state.rows,
        [],
        ImportBatch(
            imported_at=datetime.now(timezone.utc),
            image_count=1,
            game_version=GameVersion.RU_YUAN,
        ),
        state.catalog.order_index,
    )

    workbook = load_workbook(destination, data_only=True)
    assert workbook.sheetnames == ["背包汇总", "经验星石", "导入与校验"]
    summary = workbook["背包汇总"]
    assert [cell.value for cell in summary[1]] == [
        "游戏版本",
        "游戏账号名称",
        "大类",
        "星石名称",
        "品质",
        "等级",
        "数量",
    ]
    values = list(summary.iter_rows(min_row=2, values_only=True))
    assert [(row[2], row[3], row[5], row[4]) for row in values] == [
        ("主星", "天府", 60, "橙"),
        ("主星", "天府", 60, "橙"),
        ("主星", "天府", 60, "紫"),
        ("主星", "武曲", 60, "橙"),
        ("主星", "天府", 40, "橙"),
        ("辅星", "天府", 60, "橙"),
    ]
    assert [row[6] for row in values] == [
        "本组共 2 颗",
        None,
        "本组共 1 颗",
        "本组共 1 颗",
        "本组共 1 颗",
        "本组共 1 颗",
    ]


def _card(
    card_id: str,
    row: int,
    column: int,
    circle: tuple[int, int, int],
) -> CardCandidate:
    x, y, radius = circle
    return CardCandidate(
        card_id,
        row,
        column,
        (x - radius, y - radius, radius * 2, radius * 2),
        (0.0, 0.0, 0.0, 0.0),
        True,
        0.97,
        (x - radius, y + radius, radius * 2, radius),
        (x, y - radius, radius, radius),
        circle,
    )


@pytest.mark.parametrize(
    ("circle", "expected"),
    [
        ((20, 5, 6), {"card": "top"}),
        ((20, 195, 6), {"card": "bottom"}),
    ],
)
def test_physical_top_and_bottom_disc_cuts_are_excluded(
    circle: tuple[int, int, int],
    expected: dict[str, str],
) -> None:
    assert LocalOfflineVisionPipeline._auto_excluded_edge_fragments(
        [_card("card", 0, 0, circle)],
        200,
    ) == expected


@pytest.mark.parametrize(
    ("edge", "circles", "boundaries"),
    [
        (
            "top",
            [(30 + column * 50, 31 + column, 12) for column in range(4)],
            {"content_top": 20},
        ),
        (
            "bottom",
            [(30 + column * 50, 169 - column, 12) for column in range(4)],
            {"content_bottom": 180},
        ),
    ],
)
def test_four_card_boundary_row_is_consistent_and_adds_zero_complete_rows(
    edge: str,
    circles: list[tuple[int, int, int]],
    boundaries: dict[str, int],
) -> None:
    cards = [
        _card(f"edge-{column}", 0, column, circle)
        for column, circle in enumerate(circles)
    ]
    excluded = LocalOfflineVisionPipeline._auto_excluded_edge_fragments(
        cards,
        200,
        **boundaries,
    )
    assert excluded == {f"edge-{column}": edge for column in range(4)}

    state = SessionState(load_catalog())
    state.detected_items = [
        DetectedStarItem(
            card_id=card.card_id,
            source_image="edge-row",
            source_position=f"r1c{card.column_index + 1}",
            page_type="main",
            final_name="天府",
            final_level=60,
            final_quality=Quality.ORANGE,
            is_complete_card=True,
            inventory_action="auto_excluded_edge_fragment",
        )
        for card in cards
    ]
    state.recalculate_postprocess()
    assert state.rows == []


def test_pending_fragment_waits_for_confirmation_and_manual_keep_can_count() -> None:
    state = SessionState(load_catalog())
    state.detected_items = [
        DetectedStarItem(
            card_id="pending",
            source_image="partial",
            final_name="天府",
            final_level=60,
            final_quality=Quality.ORANGE,
            is_complete_card=False,
            inventory_action="keep",
        ),
        DetectedStarItem(
            card_id="excluded",
            source_image="partial",
            final_name="武曲",
            final_level=50,
            final_quality=Quality.PURPLE,
            is_complete_card=True,
            inventory_action="auto_excluded_edge_fragment",
        ),
    ]
    state.recalculate_postprocess()
    assert state.rows == []
    assert review_counts(state.detected_items, len(state.rows))["fully_resolved_count"] == 0

    state.set_card_inventory_action("excluded", "keep")
    assert [row.name for row in state.rows] == ["武曲"]


def test_normal_four_card_and_terminal_one_to_three_rows_are_not_propagated() -> None:
    normal = [
        _card(f"normal-{column}", 0, column, (30 + column * 50, 80, 12))
        for column in range(4)
    ]
    assert LocalOfflineVisionPipeline._auto_excluded_edge_fragments(
        normal,
        200,
        content_top=20,
    ) == {}

    for size in (1, 2, 3):
        terminal = [
            _card(
                f"tail-{size}-{column}",
                1,
                column,
                (30 + column * 50, 31 + column * 2, 12),
            )
            for column in range(size)
        ]
        excluded = LocalOfflineVisionPipeline._auto_excluded_edge_fragments(
            terminal,
            200,
            content_top=20,
        )
        assert excluded == {f"tail-{size}-0": "top"}
