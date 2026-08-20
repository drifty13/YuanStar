from __future__ import annotations

from pathlib import Path
import unicodedata

from openpyxl import Workbook
from openpyxl.styles import Font

from .domain import ExperienceResource, GameVersion, ImportBatch, InventorySummaryRow, normalize_instance_rows, reconcile
from .ui_contract import inventory_display_rows


def _write_sheet_header(sheet, headers: list[str]) -> None:
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    sheet.freeze_panes = "A2"


def _display_width(value: object) -> int:
    return sum(
        2 if unicodedata.east_asian_width(character) in {"W", "F", "A"} else 1
        for character in str(value or "")
    )


def export_workbook(
    destination: Path,
    game_version: GameVersion,
    account_name: str,
    rows: list[InventorySummaryRow],
    experience_resources: list[ExperienceResource],
    batch: ImportBatch | None,
    catalog_order: dict[str, int],
    plan_targets: dict[str, int] | None = None,
) -> Path:
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "背包汇总"
    _write_sheet_header(summary, ["游戏版本", "游戏账号名称", "大类", "星石名称", "品质", "当前等级", "计划等级", "数量"])
    occurrences = normalize_instance_rows(rows, catalog_order)
    display_rows = inventory_display_rows(occurrences)
    by_id = {row.id: row for row in occurrences}
    for display in display_rows:
        row = by_id[str(display["id"])]
        summary.append([
            game_version.value,
            account_name or "",
            row.kind.value,
            row.name,
            row.quality.value,
            row.level,
            (plan_targets or {}).get(row.star_instance_id, row.level),
            display["group_quantity"],
        ])

    experience = workbook.create_sheet("经验星石")
    _write_sheet_header(experience, ["游戏版本", "游戏账号名称", "名称", "数量", "单颗经验"])
    for resource in experience_resources:
        experience.append([
            game_version.value,
            account_name or "",
            resource.name,
            resource.quantity,
            resource.experience_per_item,
        ])

    audit = workbook.create_sheet("导入与校验")
    _write_sheet_header(audit, ["导入时间", "上传图片数量", "背包当前数量", "背包容量", "主星与辅星汇总数量", "差值", "完整性状态", "备注"])
    expected = batch.bag_current_count if batch else None
    result = reconcile(rows, expected)
    audit.append([
        batch.imported_at.replace(tzinfo=None) if batch else None,
        batch.image_count if batch else 0,
        expected,
        batch.bag_capacity if batch else None,
        result.actual_count,
        result.difference,
        result.status.value,
        batch.note if batch else "尚未建立导入批次",
    ])
    for sheet in workbook.worksheets:
        sheet.auto_filter.ref = sheet.dimensions
        for column in sheet.columns:
            width = min(max(_display_width(cell.value) for cell in column) + 3, 40)
            sheet.column_dimensions[column[0].column_letter].width = width
    workbook.save(destination)
    return destination
