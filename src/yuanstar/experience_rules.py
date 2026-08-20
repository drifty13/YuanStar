"""Load the Phase 0.6 experience-star rules once per service process."""

from __future__ import annotations

from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path
from types import MappingProxyType
from typing import Mapping

from openpyxl import load_workbook

from .catalog import PROJECT_ROOT


DEFAULT_EXPERIENCE_RULES_PATH = (
    PROJECT_ROOT
    / "resources"
    / "reference"
    / "YuanStar_Phase0_6A_经验星曜规则与逐级数据.xlsx"
)
CONFIG_SHEET = "Codex配置"
INTERVAL_SHEET = "升级经验区间"
DETAIL_SHEET = "逐级经验明细"


class ExperienceRuleLoadError(ValueError):
    """A user-facing explanation of why experience calculation is unavailable."""


@dataclass(frozen=True)
class ExperienceRules:
    max_level: int
    level_exp: Mapping[int, int]
    white_exp: int
    purple_exp: int
    orange_exp: int
    round_exp_to: int
    stage_6_24_purple_yield: int
    stage_6_24_stamina_cost: int
    include_orange_in_owned_exp: bool
    assume_current_bar_progress: int
    include_breakthrough_materials: bool

    def __post_init__(self) -> None:
        object.__setattr__(self, "level_exp", MappingProxyType(dict(self.level_exp)))

    @property
    def stage_6_24_exp_yield(self) -> int:
        return self.stage_6_24_purple_yield * self.purple_exp


def _require_positive_int(value: object, *, label: str) -> int:
    if isinstance(value, bool) or not isinstance(value, int) or value <= 0:
        raise ExperienceRuleLoadError(f"{label}不是有效正整数")
    return value


def _require_int(value: object, *, label: str) -> int:
    if isinstance(value, bool) or not isinstance(value, int):
        raise ExperienceRuleLoadError(f"{label}不是有效整数")
    return value


def _sheet(workbook, name: str):
    if name not in workbook.sheetnames:
        raise ExperienceRuleLoadError(f"缺少“{name}”工作表")
    return workbook[name]


def _header_indexes(sheet, expected: tuple[str, ...], *, sheet_name: str) -> dict[str, int]:
    for row in sheet.iter_rows(values_only=True):
        headers = [str(value).strip() if value is not None else "" for value in row]
        if not any(headers):
            continue
        if not set(expected).intersection(headers):
            continue
        missing = [name for name in expected if name not in headers]
        if missing:
            raise ExperienceRuleLoadError(
                f"“{sheet_name}”工作表表头错误：缺少“{missing[0]}”列"
            )
        return {name: headers.index(name) for name in expected}
    raise ExperienceRuleLoadError(f"“{sheet_name}”工作表表头错误")


def _config_values(sheet) -> dict[str, object]:
    indexes = _header_indexes(sheet, ("配置键", "值"), sheet_name=CONFIG_SHEET)
    values: dict[str, object] = {}
    started = False
    for row in sheet.iter_rows(values_only=True):
        cells = list(row)
        key = cells[indexes["配置键"]] if indexes["配置键"] < len(cells) else None
        value = cells[indexes["值"]] if indexes["值"] < len(cells) else None
        if key == "配置键":
            started = True
            continue
        if not started or key is None:
            continue
        key_text = str(key).strip()
        if key_text:
            values[key_text] = value
    return values


def _required_config(values: Mapping[str, object], key: str) -> object:
    if key not in values:
        raise ExperienceRuleLoadError(f"“{CONFIG_SHEET}”缺少配置项“{key}”")
    return values[key]


def _expand_level_exp(sheet, *, max_level: int) -> dict[int, int]:
    expected = ("当前等级起", "当前等级止", "经验条规则", "首级经验", "递增步长")
    indexes = _header_indexes(sheet, expected, sheet_name=INTERVAL_SHEET)
    intervals: list[tuple[int, int, str, int, int]] = []
    started = False
    for row in sheet.iter_rows(values_only=True):
        cells = list(row)
        start_raw = cells[indexes["当前等级起"]] if indexes["当前等级起"] < len(cells) else None
        if start_raw == "当前等级起":
            started = True
            continue
        if not started or start_raw is None:
            continue
        end_raw = cells[indexes["当前等级止"]] if indexes["当前等级止"] < len(cells) else None
        rule_raw = cells[indexes["经验条规则"]] if indexes["经验条规则"] < len(cells) else None
        first_raw = cells[indexes["首级经验"]] if indexes["首级经验"] < len(cells) else None
        step_raw = cells[indexes["递增步长"]] if indexes["递增步长"] < len(cells) else None
        start = _require_positive_int(start_raw, label="当前等级起")
        end = _require_positive_int(end_raw, label="当前等级止")
        if end < start:
            raise ExperienceRuleLoadError(f"等级区间{start}—{end}无效")
        rule = str(rule_raw or "").strip()
        if rule not in {"等差递增", "固定值"}:
            raise ExperienceRuleLoadError(f"等级区间{start}—{end}的经验条规则无效")
        first = _require_positive_int(first_raw, label="首级经验")
        step = _require_int(step_raw, label="递增步长")
        if rule == "等差递增" and step < 0:
            raise ExperienceRuleLoadError(f"等级区间{start}—{end}的递增步长不能为负数")
        intervals.append((start, end, rule, first, step))

    level_exp: dict[int, int] = {}
    for start, end, rule, first, step in intervals:
        for level in range(start, end + 1):
            if level in level_exp:
                raise ExperienceRuleLoadError(f"等级数据区间重叠：{level}级")
            value = first + (level - start) * step if rule == "等差递增" else first
            level_exp[level] = _require_positive_int(value, label=f"{level}级经验")
    for level in range(1, max_level + 1):
        if level not in level_exp:
            raise ExperienceRuleLoadError(f"等级数据缺少{level}级")
    extra = [level for level in level_exp if level > max_level]
    if extra:
        raise ExperienceRuleLoadError(f"等级数据超出最高等级：{min(extra)}级")
    return level_exp


def _validate_detail_cache(path: Path, level_exp: Mapping[int, int]) -> None:
    """Compare only usable cached values; formula-free caches are optional."""
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        if DETAIL_SHEET not in workbook.sheetnames:
            return
        sheet = workbook[DETAIL_SHEET]
        indexes = _header_indexes(
            sheet,
            ("当前等级", "当前等级经验条上限"),
            sheet_name=DETAIL_SHEET,
        )
        started = False
        for row in sheet.iter_rows(values_only=True):
            cells = list(row)
            level = cells[indexes["当前等级"]] if indexes["当前等级"] < len(cells) else None
            value = cells[indexes["当前等级经验条上限"]] if indexes["当前等级经验条上限"] < len(cells) else None
            if level == "当前等级":
                started = True
                continue
            if not started or not isinstance(level, int) or isinstance(level, bool):
                continue
            if isinstance(value, int) and not isinstance(value, bool) and level in level_exp and value != level_exp[level]:
                raise ExperienceRuleLoadError(f"“{DETAIL_SHEET}”与区间规则不一致：{level}级")
    finally:
        workbook.close()


def load_experience_rules(path: Path = DEFAULT_EXPERIENCE_RULES_PATH) -> ExperienceRules:
    """Read and validate a rules workbook. This function intentionally is not cached."""
    path = Path(path)
    if not path.is_file():
        raise ExperienceRuleLoadError("经验星曜规则文件不存在")
    try:
        workbook = load_workbook(path, read_only=True, data_only=False, keep_links=True)
    except Exception as error:
        raise ExperienceRuleLoadError(f"无法读取经验星曜规则文件：{error}") from error
    try:
        if workbook._external_links:
            raise ExperienceRuleLoadError("经验星曜规则文件不得引用外部工作簿")
        config = _config_values(_sheet(workbook, CONFIG_SHEET))
        max_level = _require_positive_int(_required_config(config, "max_level"), label="最高等级")
        if max_level != 60:
            raise ExperienceRuleLoadError("最高等级必须为60")
        white_exp = _require_positive_int(_required_config(config, "white_exp"), label="白星曜经验值")
        purple_exp = _require_positive_int(_required_config(config, "purple_exp"), label="紫星曜经验值")
        orange_exp = _require_positive_int(_required_config(config, "orange_exp"), label="橙星曜经验值")
        round_exp_to = _require_positive_int(_required_config(config, "round_exp_to"), label="经验取整粒度")
        if round_exp_to != 100:
            raise ExperienceRuleLoadError("经验取整粒度必须为100")
        stage_yield = _require_positive_int(
            _required_config(config, "stage_6_24_purple_yield"), label="6-24紫星曜产出"
        )
        stamina_cost = _require_positive_int(
            _required_config(config, "stage_6_24_stamina_cost"), label="6-24体力"
        )
        include_orange = _required_config(config, "include_orange_in_owned_exp")
        current_progress = _require_int(
            _required_config(config, "assume_current_bar_progress"), label="当前经验条进度"
        )
        include_breakthrough = _required_config(config, "include_breakthrough_materials")
        if not isinstance(include_orange, bool):
            raise ExperienceRuleLoadError("是否计入橙星曜库存不是有效布尔值")
        if current_progress != 0:
            raise ExperienceRuleLoadError("当前经验条进度必须为0")
        if not isinstance(include_breakthrough, bool) or include_breakthrough:
            raise ExperienceRuleLoadError("突破材料必须不纳入经验计算")
        level_exp = _expand_level_exp(_sheet(workbook, INTERVAL_SHEET), max_level=max_level)
    finally:
        workbook.close()
    _validate_detail_cache(path, level_exp)
    return ExperienceRules(
        max_level=max_level,
        level_exp=level_exp,
        white_exp=white_exp,
        purple_exp=purple_exp,
        orange_exp=orange_exp,
        round_exp_to=round_exp_to,
        stage_6_24_purple_yield=stage_yield,
        stage_6_24_stamina_cost=stamina_cost,
        include_orange_in_owned_exp=include_orange,
        assume_current_bar_progress=current_progress,
        include_breakthrough_materials=include_breakthrough,
    )


@lru_cache(maxsize=1)
def cached_experience_rules() -> ExperienceRules:
    """The one immutable rules object shared by every page refresh in this process."""
    return load_experience_rules(DEFAULT_EXPERIENCE_RULES_PATH)
