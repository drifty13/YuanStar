from __future__ import annotations

import json
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from .domain import StarKind


PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_CATALOG_PATH = PROJECT_ROOT / "data" / "star_catalog.json"
DEFAULT_ALIASES_PATH = PROJECT_ROOT / "data" / "ocr_aliases.json"
DEFAULT_REFERENCE_PATH = PROJECT_ROOT / "resources" / "reference" / "如鸢星石表格.xlsx"
REFERENCE_SHEET_NAME = "表格视图"


@dataclass(frozen=True)
class CatalogEntry:
    name: str
    kind: StarKind
    aliases: tuple[str, ...]
    display_group: str | None
    usage_tags: tuple[str, ...]
    raw_effect_text: str | None
    experience_per_item: int | None = None
    star_description: str | None = None


class StarCatalog:
    def __init__(self, entries: list[CatalogEntry], aliases: dict[str, str]) -> None:
        self.entries = entries
        self.aliases = aliases
        self._by_name = {entry.name: entry for entry in entries}
        if len(self._by_name) != len(entries):
            raise ValueError("标准星石名称必须唯一")
        for alias, standard_name in aliases.items():
            if standard_name not in self._by_name and alias != "星耀":
                raise ValueError(f"别名 {alias} 指向未知标准名称 {standard_name}")

    def normalize(self, value: str) -> str:
        return self.aliases.get(value.strip(), value.strip())

    def by_kind(self, kind: StarKind) -> list[CatalogEntry]:
        return [entry for entry in self.entries if entry.kind == kind]

    def names_for_kind(self, kind: StarKind) -> list[str]:
        return [entry.name for entry in self.by_kind(kind)]

    def entry(self, name: str) -> CatalogEntry:
        return self._by_name[self.normalize(name)]

    def description(self, name: str) -> str | None:
        """Return the reference-sheet description for a standard name or alias."""
        return self.entry(name).star_description

    @property
    def order_index(self) -> dict[str, int]:
        return {entry.name: index for index, entry in enumerate(self.entries) if entry.kind != StarKind.EXPERIENCE}


def load_catalog(
    catalog_path: Path = DEFAULT_CATALOG_PATH,
    aliases_path: Path = DEFAULT_ALIASES_PATH,
    reference_path: Path = DEFAULT_REFERENCE_PATH,
) -> StarCatalog:
    raw_catalog = json.loads(catalog_path.read_text(encoding="utf-8"))
    aliases = json.loads(aliases_path.read_text(encoding="utf-8"))
    combined_aliases = dict(aliases)
    for item in raw_catalog["stars"]:
        for alias in item.get("aliases", []):
            name = item["name"]
            if alias in combined_aliases and combined_aliases[alias] != name:
                raise ValueError(f"别名 {alias} 重复且指向不一致")
            combined_aliases[alias] = name
    descriptions = _load_reference_descriptions(reference_path, combined_aliases)
    entries = [
        CatalogEntry(
            name=item["name"],
            kind=StarKind(item["kind"]),
            aliases=tuple(item.get("aliases", [])),
            display_group=item.get("display_group"),
            usage_tags=tuple(item.get("usage_tags", [])),
            raw_effect_text=item.get("raw_effect_text"),
            experience_per_item=item.get("experience_per_item"),
            star_description=descriptions.get(item["name"]),
        )
        for item in raw_catalog["stars"]
    ]
    return StarCatalog(entries, combined_aliases)


def _load_reference_descriptions(reference_path: Path, aliases: dict[str, str]) -> dict[str, str]:
    """Read only the named description column, preserving Excel newlines verbatim."""
    if not reference_path.exists():
        raise FileNotFoundError(f"星石说明参考表不存在：{reference_path}")
    workbook = load_workbook(reference_path, read_only=True, data_only=True)
    try:
        if REFERENCE_SHEET_NAME not in workbook.sheetnames:
            raise ValueError(f"星石说明参考表缺少工作表：{REFERENCE_SHEET_NAME}")
        sheet = workbook[REFERENCE_SHEET_NAME]
        headers = [str(cell.value or "").strip() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        try:
            name_index = headers.index("星石名称")
            description_index = headers.index("星石描述")
        except ValueError as error:
            raise ValueError("星石说明参考表缺少“星石名称”或“星石描述”列") from error
        descriptions: dict[str, str] = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if name_index >= len(row) or description_index >= len(row):
                continue
            raw_name = row[name_index]
            raw_description = row[description_index]
            if raw_name is None or raw_description is None:
                continue
            description = str(raw_description)
            if not description.strip():
                continue
            name = aliases.get(str(raw_name).strip(), str(raw_name).strip())
            descriptions[name] = description
        return descriptions
    finally:
        workbook.close()
